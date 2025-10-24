from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db import transaction
from django.views.decorators.http import require_http_methods, require_POST
from django.db.models import Sum, Avg, Q, Count, Max
from django.utils import timezone
from decimal import Decimal
from datetime import timedelta
import uuid
import json
from . import models
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.utils import PlotlyJSONEncoder

# 方法实现（数据库操作和页面跳转）
# ====================   默认跳转  ===========================
def index(request):
    return redirect("/manager/login")


# ====================   管理员登录  ===========================
def manager_login(request):
    # 直接访问（get请求），原地跳转
    if request.method == "GET":
        return render(request, "admin/admin.html")
    # 提交表单请求（POST）
    if request.method == "POST":
        # 1.获取请求参数
        number = request.POST.get("number")
        password = request.POST.get("password")
        # 2.将数据保存到数据库中（insert）
        manager = models.Manager.objects.filter(number=number, password=password)
        if manager.exists():
            name = manager.first().name
            request.session["name"] = name
            # Redirect to dashboard instead of book list
            return redirect("/manager/dashboard/")
        return redirect("/manager/login")


# ====================   管理员登出  ===========================
def manager_logout(request):
    """Manager logout function - renamed from logout to avoid conflicts"""
    request.session.clear()
    return render(request, "admin/admin.html")


# Keep the old logout function for backward compatibility
def logout(request):
    """Legacy logout function - redirects to manager_logout"""
    return manager_logout(request)


# ====================   一、出版社模块  ===========================
# 01添加出版社
def add_publisher(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 直接访问（get请求），跳转界面
    if request.method == "GET":
        return render(request, 'publisher/add_publisher.html', {"name": request.session["name"]})
    # 提交表单请求（POST）,处理数据库,跳转到列表页面
    if request.method == "POST":
        # 1.获取请求参数
        publisher_name = request.POST.get("publisher_name")
        publisher_address = request.POST.get("publisher_address")
        # 2.将数据保存到数据库中（insert）
        models.Publisher.objects.create(
            publisher_name=publisher_name,
            publisher_address=publisher_address,
        )
        # 3.重添加成功，返回出版社列表
        return redirect("/manager/publisher_list")


# 02查询所有出版社信息
def publisher_list(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 查询数据库中的所以信息（select * from）
    publisher_list = models.Publisher.objects.all()
    # 跳转到publisher_list页面，传入publisher_list数据
    return render(request, "publisher/publisher_list.html",
                  {"publisher_obj_list": publisher_list, "name": request.session["name"]})


# 03修改出版社信息
def edit_publisher(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 1获取到的是表单提交过来的内容（POST），获取对应的值
    if request.method == "POST":
        id = request.POST.get('id')
        publisher_name = request.POST.get("publisher_name")
        publisher_address = request.POST.get("publisher_address")
        # 2根据id去数据库查找对象（where id = id）
        publisher_obj = models.Publisher.objects.get(id=id)
        # 3修改
        publisher_obj.publisher_name = publisher_name
        publisher_obj.publisher_address = publisher_address
        # 04更新数据库（update）
        publisher_obj.save()
        # 4重定向到出版社列表
        return redirect('/manager/publisher_list/')
    # get请求跳转界面（获取原始数据）
    else:
        # 1获取id
        id = request.GET.get('id')
        # 2去数据库中查找相应的数据
        publisher_obj = models.Publisher.objects.get(id=id)
        publisher_obj_list = models.Publisher.objects.all()
        # 3返回页面
        return render(request, "publisher/edit_publisher.html",
                      {"publisher_obj": publisher_obj, "publisher_obj_list": publisher_obj_list,
                       "name": request.session["name"]})


# 04删除出版社信息
def delete_publisher(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 1获取要删除图书的id
    id = request.GET.get('id')
    # 2根据id删除数据库中的记录（delete）
    models.Publisher.objects.filter(id=id).delete()
    return redirect('/manager/publisher_list')


# ============================  二、图书模块操作   ===============================
# 01获取所有图书信息
def book_list(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 1获取图书信息(select *)
    book_obj_list = models.Book.objects.all()
    # 2将数据渲染到页面上
    return render(request, 'book/book_list.html', {'book_obj_list': book_obj_list, "name": request.session["name"]})


# 02添加图书
def add_book(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    if request.method == 'POST':
        # 1获取表单提交过来的内容
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        price = request.POST.get('price')
        inventory = request.POST.get('inventory')
        sale_num = request.POST.get('sale_num')
        publisher_id = request.POST.get('publisher_id')
        cover_image = request.FILES.get('cover_image')
        book_file = request.FILES.get('book_file')
        download_link = request.POST.get('download_link', '').strip()
        
        # 2保存到数据库（insert）
        book = models.Book.objects.create(
            name=name, 
            description=description,
            price=price, 
            inventory=inventory, 
            sale_num=sale_num,
            publisher_id=publisher_id
        )
        
        # Handle image upload
        if cover_image:
            book.cover_image = cover_image
        
        # Handle book file upload
        if book_file:
            book.book_file = book_file
        
        # Handle download link
        if download_link:
            book.download_link = download_link
            
        book.save()
        
        # 3重定向到图书列表页面
        return redirect('/manager/book_list/')
    else:
        # 1获取所有的出版社（点击添加图书按钮时，得到所有出版社信息供用户选择）
        publisher_obj_list = models.Publisher.objects.all()
        # 2返回html页面（在页面中遍历出版社对象列表）
        return render(request, 'book/add_book.html', locals())


# 03修改图书信息
def edit_book(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 点击修改图书（获取要修改图书的原本信息）
    if request.method == 'GET':
        id = request.GET.get('id')
        book_obj = models.Book.objects.get(id=id)
        publisher_obj_list = models.Publisher.objects.all()
        book_obj_list = models.Book.objects.all()
        return render(request, "book/edit_book.html",
                      {"book_obj": book_obj, "book_obj_list": book_obj_list, "publisher_obj_list": publisher_obj_list,
                       "name": request.session["name"]})
    # 修改图书信息（POST表单）
    else:
        id = request.POST.get('id')
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        inventory = request.POST.get('inventory')
        price = request.POST.get('price')
        sale_num = request.POST.get('sale_num')
        publisher_id = request.POST.get('publisher_id')
        cover_image = request.FILES.get('cover_image')
        book_file = request.FILES.get('book_file')
        download_link = request.POST.get('download_link', '').strip()
        
        # 获取要更新的图书对象
        book = models.Book.objects.get(id=id)
        
        # 数据库中更新图书信息
        book.name = name
        book.description = description
        book.inventory = inventory
        book.price = price
        book.sale_num = sale_num
        book.publisher_id = publisher_id
        
        # Handle image upload
        if cover_image:
            book.cover_image = cover_image
        
        # Handle book file upload
        if book_file:
            book.book_file = book_file
        
        # Handle download link
        if download_link:
            book.download_link = download_link
        elif 'clear_download_link' in request.POST:
            book.download_link = None
            
        book.save()
        return redirect("/manager/book_list/")


# 04删除图书
def delete_book(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    id = request.GET.get('id')
    models.Book.objects.filter(id=id).delete()
    return redirect('/manager/book_list')


# ================================  三、作者操作模块  =============================
# 01作者列表
def author_list(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 列表存储作者和图书（一个作者可能有多本书）
    ret_list = []
    # 获取作者列表信息
    author_obj_list = models.Author.objects.all()
    for author_obj in author_obj_list:
        book_obj_list = author_obj.book.all()
        # 定义字典存储
        ret_dic = {'author_obj': author_obj, 'book_list': book_obj_list}
        ret_list.append(ret_dic)
    return render(request, 'author/author_list.html', {'ret_list': ret_list, "name": request.session["name"]})


# 02添加作家
def add_author(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 添加作家页面
    if request.method == 'GET':
        # 1获取所有的图书
        book_obj_list = models.Book.objects.all()
        # 2返回页面
        return render(request, 'author/add_author.html',
                      {'book_obj_list': book_obj_list, "name": request.session["name"]})
    # 数据库处理添加
    else:
        # 1.获取表单提交过来的数据
        name = request.POST.get('name')
        book_ids = request.POST.getlist('books')
        # 2 保存数据库
        author_obj = models.Author.objects.create(name=name)  # 创建对象
        author_obj.book.set(book_ids)  # 设置关系
        # 3 重定向到列表页面
        return redirect('/manager/author_list/')


# 03修改作者信息
def edit_author(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    # 跳转修改界面
    if request.method == 'GET':
        id = request.GET.get('id')
        try:
            author_obj = models.Author.objects.get(id=id)
            book_obj_list = models.Book.objects.all()
            # Get currently associated books
            current_books = author_obj.book.all()
            return render(request, 'author/edit_author.html',
                          {
                              'author_obj': author_obj, 
                              'book_obj_list': book_obj_list,
                              'current_books': current_books,
                              "name": request.session["name"]
                          })
        except models.Author.DoesNotExist:
            return redirect('/manager/author_list/')
    
    # 提交修改表单
    else:
        id = request.POST.get('id')
        name = request.POST.get('name')
        book_ids = request.POST.getlist('books')  # Get list of selected book IDs
        
        try:
            # 找到作者对象
            author_obj = models.Author.objects.filter(id=id).first()
            if author_obj:
                author_obj.name = name
                # Clear existing relationships and set new ones
                author_obj.book.set(book_ids)  # This handles the many-to-many relationship
                author_obj.save()
            return redirect('/manager/author_list/')
        except Exception as e:
            # Handle any errors gracefully
            return redirect(f'/manager/edit_author/?id={id}')

# 04 删除作者
def delete_author(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 1获取id
    id = request.GET.get('id')
    # 2删除作者
    models.Author.objects.filter(id=id).delete()
    # 3重定向作者列表 - Fixed URL
    return redirect('/manager/author_list')


# ====================   PUBLIC USER INTERFACE  ===========================

def public_home(request):
    """Public homepage with book statistics and featured content"""
    book_count = models.Book.objects.count()
    author_count = models.Author.objects.count()
    publisher_count = models.Publisher.objects.count()
    
    # Get featured books (top 6 by sales)
    featured_books = models.Book.objects.select_related('publisher').order_by('-sale_num')[:6]
    
    # Recent books (last 8 added)
    recent_books = models.Book.objects.select_related('publisher').order_by('-id')[:8]
    
    context = {
        'book_count': book_count,
        'author_count': author_count,
        'publisher_count': publisher_count,
        'featured_books': featured_books,
        'recent_books': recent_books,
    }
    return render(request, 'public/home.html', context)

def public_books(request):
    """Public book listing with search and pagination"""
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', 'name')
    
    books = models.Book.objects.select_related('publisher')
    
    if search_query:
        books = books.filter(name__icontains=search_query)
    
    # Sorting options
    if sort_by == 'price_low':
        books = books.order_by('price')
    elif sort_by == 'price_high':
        books = books.order_by('-price')
    elif sort_by == 'popular':
        books = books.order_by('-sale_num')
    else:
        books = books.order_by('name')
    
    context = {
        'books': books,
        'search_query': search_query,
        'sort_by': sort_by,
    }
    return render(request, 'public/books.html', context)

def public_book_detail(request, book_id):
    """Public book detail view"""
    book = get_object_or_404(models.Book.objects.select_related('publisher'), id=book_id)
    authors = book.author_set.all()
    
    # Related books by same publisher
    related_books = models.Book.objects.filter(
        publisher=book.publisher
    ).exclude(id=book.id)[:4]
    
    context = {
        'book': book,
        'authors': authors,
        'related_books': related_books,
    }
    return render(request, 'public/book_detail.html', context)

def public_authors(request):
    """Public authors listing"""
    search_query = request.GET.get('search', '')
    
    # Fix: Use 'book' instead of 'book_set' for ManyToMany relationship
    authors = models.Author.objects.prefetch_related('book').all()
    
    if search_query:
        authors = authors.filter(name__icontains=search_query)
    
    authors = authors.order_by('name')
    
    context = {
        'authors': authors,
        'search_query': search_query,
    }
    return render(request, 'public/authors.html', context)

def public_author_detail(request, author_id):
    """Public author detail view"""
    author = get_object_or_404(models.Author, id=author_id)
    # Fix: Use 'book' instead of 'book_set' for ManyToMany relationship
    books = author.book.select_related('publisher').all()
    
    # Calculate statistics
    total_sales = sum(book.sale_num for book in books)
    total_inventory = sum(book.inventory for book in books)
    avg_price = sum(book.price for book in books) / len(books) if books else 0
    
    context = {
        'author': author,
        'books': books,
        'total_sales': total_sales,
        'total_inventory': total_inventory,
        'avg_price': avg_price,
    }
    return render(request, 'public/author_detail.html', context)

def public_publishers(request):
    """Public publishers listing"""
    search_query = request.GET.get('search', '')
    
    publishers = models.Publisher.objects.all()
    
    if search_query:
        publishers = publishers.filter(publisher_name__icontains=search_query)
    
    publishers = publishers.order_by('publisher_name')
    
    context = {
        'publishers': publishers,
        'search_query': search_query,
    }
    return render(request, 'public/publishers.html', context)

def public_publisher_detail(request, publisher_id):
    """Public publisher detail view"""
    publisher = get_object_or_404(models.Publisher, id=publisher_id)
    books = models.Book.objects.filter(publisher=publisher).order_by('name')
    
    # Calculate statistics
    total_sales = sum(book.sale_num for book in books)
    total_inventory = sum(book.inventory for book in books)
    avg_price = sum(book.price for book in books) / len(books) if books else 0
    
    # Count unique authors
    author_count = models.Author.objects.filter(book__publisher=publisher).distinct().count()
    
    context = {
        'publisher': publisher,
        'books': books,
        'total_sales': total_sales,
        'total_inventory': total_inventory,
        'avg_price': avg_price,
        'author_count': author_count,
    }
    return render(request, 'public/publisher_detail.html', context)

# ==================== E-COMMERCE FUNCTIONALITY ====================

def get_session_key(request):
    """Get or create session key for cart functionality"""
    if not request.session.session_key:
        request.session.create()
    return request.session.session_key

@require_POST
def add_to_cart(request, book_id):
    """Add book to shopping cart"""
    try:
        book = get_object_or_404(models.Book, id=book_id)
        quantity = int(request.POST.get('quantity', 1))
        session_key = get_session_key(request)
        
        # Validate quantity
        if quantity > book.inventory:
            return JsonResponse({
                'success': False,
                'message': f'库存不足！当前库存：{book.inventory}本'
            })
        
        # Get or create cart item
        cart_item, created = models.CartItem.objects.get_or_create(
            session_key=session_key,
            book=book,
            defaults={'quantity': quantity}
        )
        
        if not created:
            # Update existing cart item
            new_quantity = cart_item.quantity + quantity
            if new_quantity > book.inventory:
                return JsonResponse({
                    'success': False,
                    'message': f'库存不足！当前库存：{book.inventory}本，购物车中已有：{cart_item.quantity}本'
                })
            cart_item.quantity = new_quantity
            cart_item.save()
        
        # Get total cart count
        cart_count = models.CartItem.objects.filter(session_key=session_key).count()
        
        return JsonResponse({
            'success': True,
            'message': f'已将《{book.name}》添加到购物车',
            'cart_count': cart_count,
            'item_quantity': cart_item.quantity
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': '添加到购物车失败，请重试'
        })

def view_cart(request):
    """Display shopping cart"""
    session_key = get_session_key(request)
    cart_items = models.CartItem.objects.filter(session_key=session_key).select_related('book')
    
    total_amount = sum(item.get_total_price() for item in cart_items)
    total_items = sum(item.quantity for item in cart_items)
    
    context = {
        'cart_items': cart_items,
        'total_amount': total_amount,
        'total_items': total_items,
    }
    
    return render(request, 'public/cart.html', context)

@require_POST
def update_cart(request):
    """Update cart item quantities via AJAX"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        quantity = int(data.get('quantity', 1))
        session_key = get_session_key(request)
        
        cart_item = get_object_or_404(models.CartItem, id=item_id, session_key=session_key)
        
        # Validate quantity
        if quantity > cart_item.book.inventory:
            return JsonResponse({
                'success': False,
                'message': f'库存不足！最大可购买：{cart_item.book.inventory}本'
            })
        
        if quantity <= 0:
            cart_item.delete()
            message = f'已从购物车移除《{cart_item.book.name}》'
        else:
            cart_item.quantity = quantity
            cart_item.save()
            message = f'已更新《{cart_item.book.name}》数量为：{quantity}本'
        
        # Recalculate totals
        cart_items = models.CartItem.objects.filter(session_key=session_key)
        total_amount = sum(item.get_total_price() for item in cart_items)
        total_items = sum(item.quantity for item in cart_items)
        
        return JsonResponse({
            'success': True,
            'message': message,
            'item_total': float(cart_item.get_total_price()) if quantity > 0 else 0,
            'cart_total': float(total_amount),
            'total_items': total_items
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': '更新购物车失败'
        })

def remove_from_cart(request, item_id):
    """Remove item from cart"""
    session_key = get_session_key(request)
    cart_item = get_object_or_404(models.CartItem, id=item_id, session_key=session_key)
    book_name = cart_item.book.name
    cart_item.delete()
    
    messages.success(request, f'已从购物车移除《{book_name}》')
    return redirect('manager:view_cart')

def get_cart_count(request):
    """Get cart item count for AJAX requests"""
    session_key = get_session_key(request)
    cart_count = models.CartItem.objects.filter(session_key=session_key).count()
    return JsonResponse({'cart_count': cart_count})

@require_POST
def buy_now(request, book_id):
    """Direct purchase without cart"""
    try:
        book = get_object_or_404(models.Book, id=book_id)
        quantity = int(request.POST.get('quantity', 1))
        session_key = get_session_key(request)
        
        # Validate quantity
        if quantity > book.inventory:
            messages.error(request, f'库存不足！当前库存：{book.inventory}本')
            return redirect('manager:public_book_detail', book_id=book_id)
        
        # Clear any existing cart items for this session (optional)
        models.CartItem.objects.filter(session_key=session_key).delete()
        
        # Add to cart for checkout
        models.CartItem.objects.create(
            session_key=session_key,
            book=book,
            quantity=quantity
        )
        
        # Redirect to checkout
        return redirect('manager:checkout')
        
    except Exception as e:
        messages.error(request, '购买失败，请重试')
        return redirect('manager:public_book_detail', book_id=book_id)

def checkout(request):
    """Checkout process - Fixed price ¥6.99 per book"""
    session_key = get_session_key(request)
    cart_items = models.CartItem.objects.filter(session_key=session_key).select_related('book')
    
    if not cart_items.exists():
        messages.warning(request, '购物车为空，请先添加商品')
        return redirect('manager:public_books')
    
    # Fixed price per book
    FIXED_PRICE = Decimal('6.99')
    
    if request.method == 'POST':
        try:
            # Calculate total with fixed price
            total_items_count = sum(item.quantity for item in cart_items)
            total_amount = FIXED_PRICE * total_items_count
            
            # Determine initial status based on payment confirmation
            payment_confirmed = request.POST.get('payment_confirmed', 'no')
            if payment_confirmed == 'yes':
                initial_status = 'processing'  # 处理中
                payment_status = 'pending'
            else:
                initial_status = 'payment_pending'  # 待付款
                payment_status = 'pending'
            
            # Create order (Digital Product - no shipping address needed)
            order = models.Order.objects.create(
                customer_name=request.POST.get('customer_name'),
                customer_email=request.POST.get('customer_email'),
                customer_phone=request.POST.get('customer_phone'),
                country=request.POST.get('country', 'China'),
                payment_method=request.POST.get('payment_method'),
                total_amount=total_amount,
                status=initial_status,
                payment_status=payment_status,
                customer_notes=request.POST.get('customer_notes', '')
            )
            
            # Create order items and update inventory
            for cart_item in cart_items:
                models.OrderItem.objects.create(
                    order=order,
                    book=cart_item.book,
                    quantity=cart_item.quantity,
                    unit_price=cart_item.book.price,
                    total_price=cart_item.get_total_price()
                )
                
                # Update book inventory and sales
                book = cart_item.book
                book.inventory -= cart_item.quantity
                book.sale_num += cart_item.quantity
                book.save()
            
            # Clear cart
            cart_items.delete()
            
            return redirect('manager:order_confirmation', order_number=order.order_number)
            
        except Exception as e:
            messages.error(request, '订单创建失败，请重试')
    
    # Calculate with fixed price
    total_items = sum(item.quantity for item in cart_items)
    total_amount = FIXED_PRICE * total_items
    
    # Payment methods grouped by region
    payment_methods_by_region = {
        'africa': [
            ('mtn_money', 'MTN Money'),
            ('orange_money', 'Orange Money'),
            ('airtel_money', 'Airtel Money'),
        ],
        'china': [
            ('wechat_pay', '微信支付'),
            ('alipay', '支付宝'),
        ],
        'others': [
            ('paypal', 'PayPal'),
            ('credit_card', '信用卡'),
            ('bank_transfer', '银行转账'),
        ]
    }
    
    context = {
        'cart_items': cart_items,
        'total_amount': total_amount,
        'total_items': total_items,
        'fixed_price': FIXED_PRICE,
        'payment_methods_by_region': payment_methods_by_region,
    }
    
    return render(request, 'public/checkout.html', context)

def order_confirmation(request, order_number):
    """Order confirmation page"""
    order = get_object_or_404(models.Order, order_number=order_number)
    order_items = models.OrderItem.objects.filter(order=order).select_related('book')
    
    context = {
        'order': order,
        'order_items': order_items,
    }
    
    return render(request, 'public/order_confirmation.html', context)

def track_order(request):
    """Order tracking page - Search by order number or email"""
    order = None
    orders = None
    
    if request.method == 'POST':
        search_type = request.POST.get('search_type', 'order_number')
        
        if search_type == 'email':
            # Search by email - return all orders
            customer_email = request.POST.get('customer_email')
            if customer_email:
                orders = models.Order.objects.filter(
                    customer_email=customer_email
                ).order_by('-created_at')
                
                if not orders.exists():
                    messages.error(request, f'未找到与邮箱 {customer_email} 相关的订单')
                    orders = None
        else:
            # Search by order number - return single order
            order_number = request.POST.get('order_number')
            if order_number:
                try:
                    order = models.Order.objects.get(order_number=order_number)
                    # Check if payment window expired and auto-cancel
                    if order.status == 'payment_pending':
                        order.auto_cancel_if_expired()
                except models.Order.DoesNotExist:
                    messages.error(request, '订单号不存在')
    
    context = {
        'order': order,
        'orders': orders
    }
    return render(request, 'public/track_order.html', context)

def download_book(request, order_id, book_id):
    """Download purchased ebook - supports files and external links"""
    from django.http import FileResponse, HttpResponse
    import os
    
    # Verify order and book
    order = get_object_or_404(models.Order, id=order_id)
    book = get_object_or_404(models.Book, id=book_id)
    
    # Check if order status allows download (shipped or delivered means payment is done)
    valid_statuses = ['paid', 'confirmed', 'processing', 'shipped', 'delivered']
    if order.status not in valid_statuses:
        messages.error(request, '订单尚未完成支付，无法下载')
        return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # Additional check: if status is shipped/delivered, payment must be completed
    # (For backwards compatibility, we allow download if status is shipped regardless of payment_status)
    if order.status not in ['shipped', 'delivered'] and order.payment_status != 'completed':
        messages.error(request, '订单支付未完成，无法下载')
        return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # Check if this book is in the order
    order_item = models.OrderItem.objects.filter(order=order, book=book).first()
    if not order_item:
        messages.error(request, '此订单不包含该图书')
        return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # Check if book has download available
    if not book.has_download():
        messages.error(request, '该图书暂无下载文件')
        return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # If book has a file, serve it directly
    if book.book_file:
        try:
            response = FileResponse(book.book_file.open('rb'))
            file_name = os.path.basename(book.book_file.name)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
        except Exception as e:
            messages.error(request, f'文件下载失败: {str(e)}')
            return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # If book has external download link, redirect to it
    elif book.download_link:
        return redirect(book.download_link)
    
    # Fallback: Create a placeholder PDF with book information
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from io import BytesIO
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        # Add content to PDF
        p.setFont("Helvetica-Bold", 24)
        p.drawString(100, 750, f"{book.name}")
        
        p.setFont("Helvetica", 12)
        p.drawString(100, 720, f"Publisher: {book.publisher.publisher_name}")
        p.drawString(100, 700, f"Order Number: {order.order_number}")
        p.drawString(100, 680, f"Customer: {order.customer_name}")
        
        p.setFont("Helvetica-Oblique", 10)
        p.drawString(100, 650, "Thank you for your purchase!")
        p.drawString(100, 630, "This is a sample ebook file.")
        p.drawString(100, 610, "Please contact support to get the actual ebook file.")
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        
        response = FileResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{book.name}.pdf"'
        
        return response
        
    except ImportError:
        # If reportlab is not installed, return a simple text file
        content = f"""
        电子书信息
        ==========
        
        书名: {book.name}
        出版社: {book.publisher.publisher_name}
        订单号: {order.order_number}
        客户: {order.customer_name}
        
        感谢您的购买！
        
        注意: 这是一个示例文件。实际生产环境中，这里应该提供真实的电子书文件。
        """
        
        response = HttpResponse(content, content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{book.name}.txt"'
        return response


# ====================   API Views for Order Actions  ===========================
def api_cancel_order(request):
    """API endpoint to cancel an order"""
    from django.http import JsonResponse
    import json
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_number = data.get('order_number')
            
            order = models.Order.objects.get(order_number=order_number)
            
            # Only allow cancellation for pending or payment_pending orders
            if order.status in ['pending', 'payment_pending']:
                order.status = 'cancelled'
                order.save()
                return JsonResponse({'success': True, 'message': '订单已取消'})
            else:
                return JsonResponse({'success': False, 'message': '此订单状态无法取消'})
                
        except models.Order.DoesNotExist:
            return JsonResponse({'success': False, 'message': '订单不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': '无效请求'})


def api_confirm_payment(request):
    """API endpoint to confirm payment"""
    from django.http import JsonResponse
    import json
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_number = data.get('order_number')
            
            order = models.Order.objects.get(order_number=order_number)
            
            # Update order status to processing
            if order.status == 'payment_pending':
                order.status = 'processing'
                order.payment_status = 'pending'
                order.save()
                return JsonResponse({'success': True, 'message': '支付确认成功'})
            else:
                return JsonResponse({'success': False, 'message': '订单状态无法确认支付'})
                
        except models.Order.DoesNotExist:
            return JsonResponse({'success': False, 'message': '订单不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': '无效请求'})


# ====================   订单管理模块  ===========================
def order_list(request):
    """Admin order list view"""
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    payment_status_filter = request.GET.get('payment_status', '')
    search_query = request.GET.get('search', '')
    
    # Start with all orders
    orders = models.Order.objects.all().select_related().prefetch_related('orderitem_set__book')
    
    # Apply filters
    if status_filter:
        orders = orders.filter(status=status_filter)
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    if search_query:
        orders = orders.filter(
            models.Q(order_number__icontains=search_query) |
            models.Q(customer_name__icontains=search_query) |
            models.Q(customer_email__icontains=search_query) |
            models.Q(customer_phone__icontains=search_query)
        )
      # Order by creation date (newest first)
    orders = orders.order_by('-created_at')
    
    # Get statistics
    total_orders = models.Order.objects.count()
    pending_orders = models.Order.objects.filter(status='pending').count()
    completed_orders = models.Order.objects.filter(payment_status='completed').count()
    total_revenue = models.Order.objects.filter(payment_status='completed').aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    
    context = {
        'orders': orders,
        'name': request.session["name"],
        'status_choices': models.ORDER_STATUS_CHOICES,
        'payment_status_choices': models.PAYMENT_STATUS_CHOICES,
        'current_status_filter': status_filter,
        'current_payment_status_filter': payment_status_filter,
        'current_search': search_query,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'completed_orders': completed_orders,
        'total_revenue': total_revenue,
    }
    
    return render(request, 'order/order_list.html', context)


def order_detail(request, order_id):
    """Admin order detail view"""
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    order = get_object_or_404(models.Order, id=order_id)
    order_items = models.OrderItem.objects.filter(order=order).select_related('book')
    
    context = {
        'order': order,
        'order_items': order_items,
        'name': request.session["name"],
        'status_choices': models.ORDER_STATUS_CHOICES,
        'payment_status_choices': models.PAYMENT_STATUS_CHOICES,
    }
    
    return render(request, 'order/order_detail.html', context)


@require_POST
def update_order_status(request):
    """Update order status via AJAX"""
    # 登录判断
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '请先登录'})
    
    try:
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('status')
        admin_notes = request.POST.get('admin_notes', '')
        
        order = get_object_or_404(models.Order, id=order_id)
        old_status = order.status
        
        order.status = new_status
        if admin_notes:
            order.admin_notes = admin_notes
        order.save()
        
        # Log the status change
        return JsonResponse({
            'success': True, 
            'message': f'订单状态已从 "{dict(models.ORDER_STATUS_CHOICES)[old_status]}" 更新为 "{dict(models.ORDER_STATUS_CHOICES)[new_status]}"',
            'new_status': new_status,
            'new_status_display': dict(models.ORDER_STATUS_CHOICES)[new_status],
            'new_status_color': order.get_status_color()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': '更新失败，请重试'})


@require_POST
def update_payment_status(request):
    """Update payment status via AJAX"""
    # 登录判断
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '请先登录'})
    
    try:
        order_id = request.POST.get('order_id')
        new_payment_status = request.POST.get('payment_status')
        transaction_id = request.POST.get('transaction_id', '')
        
        order = get_object_or_404(models.Order, id=order_id)
        old_payment_status = order.payment_status
        
        order.payment_status = new_payment_status
        if transaction_id:
            order.payment_transaction_id = transaction_id
        
        # If payment is completed, also update payment completion time
        if new_payment_status == 'completed':
            order.payment_completed_at = timezone.now()
            # Also update order status if it's still pending
            if order.status == 'pending':
                order.status = 'paid'
        
        order.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'支付状态已从 "{dict(models.PAYMENT_STATUS_CHOICES)[old_payment_status]}" 更新为 "{dict(models.PAYMENT_STATUS_CHOICES)[new_payment_status]}"',
            'new_payment_status': new_payment_status,
            'new_payment_status_display': dict(models.PAYMENT_STATUS_CHOICES)[new_payment_status],
            'new_payment_status_color': order.get_payment_status_color(),
            'order_status': order.status,
            'order_status_display': dict(models.ORDER_STATUS_CHOICES)[order.status],
            'order_status_color': order.get_status_color()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': '更新失败，请重试'})

@require_http_methods(["GET"])
def export_orders(request):
    """Export orders to CSV or Excel format"""
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    export_format = request.GET.get('format', 'csv')  # csv or excel
    status_filter = request.GET.get('status', '')
    payment_status_filter = request.GET.get('payment_status', '')
    search_filter = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Build queryset with filters
    orders = models.Order.objects.all().order_by('-created_at')
    
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    if payment_status_filter:
        orders = orders.filter(payment_status=status_filter)
    
    if search_filter:
        orders = orders.filter(
            models.Q(order_number__icontains=search_filter) |
            models.Q(customer_name__icontains=search_filter) |
            models.Q(customer_email__icontains=search_filter) |
            models.Q(customer_phone__icontains=search_filter)
        )
    
    if date_from:
        try:
            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=to_date)
        except ValueError:
            pass
    
    if export_format == 'excel':
        return export_orders_excel(orders)
    else:
        return export_orders_csv(orders)


def export_orders_csv(orders):
    """Export orders to CSV format"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="orders_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Add BOM for UTF-8 to ensure proper encoding in Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        '订单号', '客户姓名', '客户邮箱', '客户电话', '收货地址', '城市', '省份',
        '支付方式', '订单状态', '支付状态', '总金额', '商品数量', '创建时间', '客户备注', '管理员备注'
    ])
    
    # Write data
    for order in orders:
        writer.writerow([
            order.order_number,
            order.customer_name,
            order.customer_email,
            order.customer_phone,
            order.shipping_address,
            order.shipping_city,
            order.shipping_state,
            order.get_payment_method_display(),
            order.get_status_display(),
            order.get_payment_status_display(),
            f'¥{order.total_amount}',
            order.items.count(),
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.customer_notes or '',
            order.admin_notes or ''
        ])
    
    return response


def export_orders_excel(orders):
    """Export orders to Excel format"""
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "订单导出"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        '订单号', '客户姓名', '客户邮箱', '客户电话', '收货地址', '城市', '省份',
        '支付方式', '订单状态', '支付状态', '总金额', '商品数量', '创建时间', '客户备注', '管理员备注'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_num, order in enumerate(orders, 2):
        data = [
            order.order_number,
            order.customer_name,
            order.customer_email,
            order.customer_phone,
            order.shipping_address,
            order.shipping_city,
            order.shipping_state,
            order.get_payment_method_display(),
            order.get_status_display(),
            order.get_payment_status_display(),
            float(order.total_amount),
            order.items.count(),
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.customer_notes or '',
            order.admin_notes or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="orders_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@require_http_methods(["GET"])
def order_analytics(request):
    """Get order analytics data"""
    if not request.session.get('is_login'):
        return JsonResponse({'error': '请先登录'}, status=401)
    
    # Date range for analytics (default: last 30 days)
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get date range from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        try:
            start_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_to:
        try:
            end_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Analytics queries
    orders_in_range = models.Order.objects.filter(
        created_at__date__range=[start_date, end_date]
    )
    
    # Daily order counts
    daily_orders = []
    current_date = start_date
    while current_date <= end_date:
        count = orders_in_range.filter(created_at__date=current_date).count()
        revenue = orders_in_range.filter(
            created_at__date=current_date,
            payment_status='paid'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        daily_orders.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'orders': count,
            'revenue': float(revenue)
        })
        current_date += timedelta(days=1)
    
    # Status distribution
    status_counts = {}
    for status, label in models.ORDER_STATUS_CHOICES:
        count = orders_in_range.filter(status=status).count()
        status_counts[label] = count
    
    # Payment status distribution
    payment_status_counts = {}
    for status, label in models.PAYMENT_STATUS_CHOICES:
        count = orders_in_range.filter(payment_status=status).count()
        payment_status_counts[label] = count
    
    # Top customers
    top_customers = orders_in_range.values('customer_name', 'customer_email').annotate(
        order_count=models.Count('id'),
        total_spent=Sum('total_amount')
    ).order_by('-total_spent')[:10]
    
    return JsonResponse({
        'success': True,
        'data': {
            'date_range': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d')
            },
            'daily_orders': daily_orders,
            'status_distribution': status_counts,
            'payment_status_distribution': payment_status_counts,
            'top_customers': list(top_customers),
            'summary': {
                'total_orders': orders_in_range.count(),                'total_revenue': float(orders_in_range.filter(payment_status='paid').aggregate(
                    total=Sum('total_amount'))['total'] or 0),
                'average_order_value': float(orders_in_range.aggregate(
                    avg=Avg('total_amount'))['avg'] or 0),
                'completion_rate': round(
                    (orders_in_range.filter(status='delivered').count() / 
                     max(orders_in_range.count(), 1)) * 100, 2
                )
            }
        }
    })

# ====================   MANAGER DASHBOARD  ===========================
def manager_dashboard(request):
    """Professional dashboard for managers with statistics and analytics"""
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Sum, Count, Q
    import json
    
    # Get current date ranges
    now = timezone.now()
    current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month = (current_month - timedelta(days=1)).replace(day=1)
    last_7_days = now.date() - timedelta(days=7)
    
    # ==== BASIC STATISTICS ====
    total_books = models.Book.objects.count()
    total_publishers = models.Publisher.objects.count()
    total_authors = models.Author.objects.count()
    
    # Books added this month - Real dynamic calculation
    new_books_this_month = models.Book.objects.filter(
        created_at__gte=current_month
    ).count() if hasattr(models.Book, 'created_at') else 0
    
    # ==== ORDER STATISTICS ====
    try:
        total_orders = models.Order.objects.count()
        orders_this_month = models.Order.objects.filter(created_at__gte=current_month).count()
        
        # Order status counts
        pending_orders = models.Order.objects.filter(status='payment_pending').count()
        processing_orders = models.Order.objects.filter(status='processing').count()
        shipped_orders = models.Order.objects.filter(status='shipped').count()
        completed_orders = models.Order.objects.filter(status='delivered').count()
        cancelled_orders = models.Order.objects.filter(status='cancelled').count()
        
        # Revenue calculations
        total_revenue = models.Order.objects.filter(
            payment_status__in=['completed', 'pending']
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        revenue_this_month = models.Order.objects.filter(
            created_at__gte=current_month,
            payment_status__in=['completed', 'pending']
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
    except Exception as e:
        total_orders = 0
        orders_this_month = 0
        pending_orders = 0
        processing_orders = 0
        shipped_orders = 0
        completed_orders = 0
        cancelled_orders = 0
        total_revenue = 0
        revenue_this_month = 0
    
    # ==== INVENTORY STATISTICS ====
    low_inventory_books = models.Book.objects.filter(inventory__lt=10).count()
    total_inventory = models.Book.objects.aggregate(Sum('inventory'))['inventory__sum'] or 0
    total_sales = models.Book.objects.aggregate(Sum('sale_num'))['sale_num__sum'] or 0
    
    # Top selling books
    top_books = models.Book.objects.order_by('-sale_num')[:5]
    
    # Recent orders (if available)
    try:
        recent_orders = models.Order.objects.order_by('-created_at')[:5]
    except:
        recent_orders = []
    
    # ==== CHART DATA ==== 
    # Daily sales for the last 7 days
    daily_sales = []
    for i in range(6, -1, -1):  # Reverse order to show oldest to newest
        date = (now.date() - timedelta(days=i))
        try:
            day_orders = models.Order.objects.filter(
                created_at__date=date
            ).count()
            day_revenue = models.Order.objects.filter(
                created_at__date=date,
                payment_status__in=['completed', 'pending']
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        except:
            day_orders = 0
            day_revenue = 0
        
        daily_sales.append({
            'date': date.strftime('%m-%d'),
            'orders': day_orders,
            'revenue': float(day_revenue)
        })
    
    # Publisher distribution
    publisher_stats = []
    for publisher in models.Publisher.objects.all()[:5]:
        book_count = models.Book.objects.filter(publisher=publisher).count()
        publisher_stats.append({
            'name': publisher.publisher_name,
            'books': book_count
        })
    
    # Add default data if no publishers
    if not publisher_stats:
        publisher_stats = [{'name': '暂无数据', 'books': 1}]
    
    context = {
        'name': request.session["name"],
        'current_date': now.strftime('%Y年%m月%d日'),
        
        # Basic stats
        'total_books': total_books,
        'total_publishers': total_publishers,
        'total_authors': total_authors,
        'new_books_this_month': new_books_this_month,
        'total_customers': 0,  # Default value
        'new_customers_this_month': 0,  # Default value
        
        # Order stats
        'total_orders': total_orders,
        'orders_this_month': orders_this_month,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'completed_orders': completed_orders,
        'cancelled_orders': cancelled_orders,
        'total_revenue': total_revenue,
        'revenue_this_month': revenue_this_month,
        
        # Inventory stats
        'low_inventory_books': low_inventory_books,
        'total_inventory': total_inventory,
        'total_sales': total_sales,
        
        # Lists
        'top_books': top_books,
        'recent_orders': recent_orders,
        'recent_activities': [],  # Default empty list
        
        # Chart data (as JSON)
        'daily_sales_json': json.dumps(daily_sales),
        'publisher_stats_json': json.dumps(publisher_stats),
    }
    
    return render(request, 'manager/dashboard.html', context)

@require_http_methods(["GET"])
def dashboard_analytics_api(request):
    """API endpoint for dashboard analytics data"""
    if "name" not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    # Return analytics data for AJAX requests
    return JsonResponse({
        'success': True,
        'data': {
            'total_books': models.Book.objects.count(),
            'total_orders': models.Order.objects.count() if hasattr(models, 'Order') else 0,
            'low_inventory': models.Book.objects.filter(inventory__lt=10).count(),
        }
    })

@require_POST
def delete_order(request, order_id):
    """Delete an order via AJAX"""
    # 登录判断
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '请先登录'})
    
    try:
        order = get_object_or_404(models.Order, id=order_id)
        
        # Check if order can be deleted (business rules)
        if order.status == 'delivered':
            return JsonResponse({
                'success': False, 
                'message': '已送达的订单不能删除'
            })
        
        if order.payment_status == 'completed':
            return JsonResponse({
                'success': False, 
                'message': '已完成支付的订单不能删除，请先处理退款'
            })
        
        # Store order info for response
        order_number = order.order_number
        customer_name = order.customer_name
        
        # Delete the order (cascade will delete related OrderItems)
        order.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'订单 {order_number} (客户: {customer_name}) 已成功删除'
        })
        
    except models.Order.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': '订单不存在'
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'删除失败：{str(e)}'
        })
